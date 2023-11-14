# VERSION 5 - 13/11/2023
# Cette application PyQt permet de comparer le contenu de deux fichiers PDF sélectionnés par l'utilisateur. Elle affiche les différences entre les
# fichiers ligne par ligne dans un tableau et met en évidence les modifications avec des couleurs. 
# Rouge : suppression
# Vert : ajout
# Orange : Texte déplacé
# Enregistrement du tablau sous format Excel OK

# Amélioration en cours : 
# Lien avec les annotations - EN COURS => format liste

import sys
import PyPDF2
import fitz
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from PyQt5.QtWidgets import (
    QApplication,
    QMainWindow,
    QFileDialog,
    QPushButton,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
    QLabel,
    QHBoxLayout,
    QLineEdit,
    QHeaderView,
    QMessageBox,
    QDialog, 
    QTextBrowser, 
    QScrollArea, 
    QSplitter, 
    QCheckBox,
    QListWidget, 
    QListWidgetItem,
    QSpacerItem, 
    QSizePolicy
)
from PyQt5.QtCore import Qt
from PyQt5.QtCore import QSize
from PyQt5.QtGui import QColor
from PyQt5.QtGui import QIcon
import difflib


class PDFComparerApp(QMainWindow):
    current_difference_index = 0
    
    def __init__(self):
        super().__init__(None)
        self.initUI()
        self.annotation_checkboxes = {}

    def initUI(self):
        self.setWindowTitle("PDF Compare & Verify - Version 5")
        self.resize(1200, 500)

        self.setWindowIcon(QIcon('comparer.png'))
    
        # QSplitter pour diviser la fenêtre en deux parties redimensionnables
        splitter = QSplitter(Qt.Horizontal)
    
        # Zone de gauche
        left_layout = QVBoxLayout()
    
        self.form_layout1 = QHBoxLayout()
        self.pdf_label1 = QLabel("Ancienne version :")
        self.select_old_button = QPushButton("Sélectionner")
        self.select_old_button.clicked.connect(self.selectOldPDF)
        self.old_pdf_entry = QLineEdit()
        self.old_pdf_entry.setPlaceholderText("Choisir un fichier PDF")
        self.form_layout1.addWidget(self.pdf_label1)
        self.form_layout1.addWidget(self.old_pdf_entry)
        self.form_layout1.addWidget(self.select_old_button)
    
        self.form_layout2 = QHBoxLayout()
        self.pdf_label2 = QLabel("Nouvelle version :")
        self.select_new_button = QPushButton("Sélectionner")
        self.select_new_button.clicked.connect(self.selectNewPDF)
        self.new_pdf_entry = QLineEdit()
        self.new_pdf_entry.setPlaceholderText("Choisir un fichier PDF")
        self.form_layout2.addWidget(self.pdf_label2)
        self.form_layout2.addWidget(self.new_pdf_entry)
        self.form_layout2.addWidget(self.select_new_button)
    
        self.compare_button = QPushButton("Comparer")
        self.compare_button.clicked.connect(self.comparePDFs)
    
        left_layout.addLayout(self.form_layout1)
        left_layout.addLayout(self.form_layout2)
        left_layout.addWidget(self.compare_button)
    
        self.table = QTableWidget()
        left_layout.addWidget(self.table)
    
        button_layout = QHBoxLayout()
    
        self.previous_diff_button = QPushButton("Différence précédente")
        self.previous_diff_button.clicked.connect(self.showPreviousDifference)
        self.next_diff_button = QPushButton("Différence suivante")
        self.next_diff_button.clicked.connect(self.showNextDifference)

        help_button = QPushButton("?")
        help_button.setMaximumWidth(25)
        help_button.clicked.connect(self.showInstructions)
        button_layout.addWidget(help_button, alignment=Qt.AlignLeft)
        export_button = QPushButton("Exporter vers Excel")
        export_button.clicked.connect(self.exportToExcel)
        button_layout.addWidget(export_button)

        # Modifiez ces lignes dans la fonction initUI de votre classe PDFComparerApp
        self.page_label = QLabel("Rechercher le texte :")
        self.page_entry = QLineEdit()
        self.page_entry.setPlaceholderText("Texte à rechercher")
        self.go_to_page_button = QPushButton("Rechercher")
        self.go_to_page_button.clicked.connect(self.goToText)

        # Ajoutez ces widgets à votre layout de boutons
        button_layout.addWidget(self.page_label)
        button_layout.addWidget(self.page_entry)
        button_layout.addWidget(self.go_to_page_button)

        left_layout.addLayout(button_layout)
    
        button_layout.addWidget(self.previous_diff_button)
        button_layout.addWidget(self.next_diff_button)
        left_layout.addLayout(button_layout)
    
        left_widget = QWidget()
        left_widget.setLayout(left_layout)
    
        # Zone de droite (zone d'annotation)
        right_layout = QVBoxLayout()
    
        # setFixedWidth pour définir la largeur de la zone de droite
        right_widget = QWidget()
        right_widget.setFixedWidth(350)
    
        # Crée un QListWidget pour afficher les annotations à droite
        self.annotation_list = QListWidget()
        self.annotation_list.setMaximumWidth(350)  # Définissez la largeur maximale
        self.annotation_list.setWordWrap(True)  # Activez le retour à la ligne
        right_layout.addWidget(self.annotation_list)
    
        right_widget.setLayout(right_layout)
    
        # Ajoute les zones gauche et droite au QSplitter
        splitter.addWidget(left_widget)
        splitter.addWidget(right_widget)
    
        central_widget = QWidget()
        central_widget.setLayout(QVBoxLayout())
        central_widget.layout().addWidget(splitter)
        self.setCentralWidget(central_widget)
    
        self.old_pdf = None
        self.new_pdf = None
        self.diff = []

    def selectOldPDF(self):
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly

        filename, _ = QFileDialog.getOpenFileName(self, "Sélectionner l'ancienne version PDF", "", "PDF Files (*.pdf);;All Files (*)", options=options)

        if filename:
            self.old_pdf = filename
            self.old_pdf_entry.setText(filename)
            self.displayOldPDFAnnotations()

    def selectNewPDF(self):
        options = QFileDialog.Options()
        options |= QFileDialog.ReadOnly

        filename, _ = QFileDialog.getOpenFileName(self, "Sélectionner la nouvelle version PDF", "", "PDF Files (*.pdf);;All Files (*)", options=options)

        if filename:
            self.new_pdf = filename
            self.new_pdf_entry.setText(filename)

    def comparePDFs(self):
        if not self.old_pdf or not self.new_pdf:
            QMessageBox.critical(self, "Erreur", "Veuillez sélectionner les deux fichiers PDF.")
            return
    
        try:
            with open(self.old_pdf, "rb") as file1, open(self.new_pdf, "rb") as file2:
                pdf1 = PyPDF2.PdfReader(file1)
                pdf2 = PyPDF2.PdfReader(file2)
    
                text1 = ""
                text2 = ""
    
                for page in range(min(len(pdf1.pages), len(pdf2.pages))):
                    text1 += pdf1.pages[page].extract_text()
                    text2 += pdf2.pages[page].extract_text()
    
                differ = difflib.Differ()
                self.diff = list(differ.compare(text1.splitlines(), text2.splitlines()))
    
                self.displayDiffResults(self.diff)  # Afficher les différences
    
        except Exception as e:
            QMessageBox.critical(self, "Erreur", f"Une erreur s'est produite : {str(e)}")

    def displayDiffResults(self, diff):
        if not self.table:
            return
    
        self.table.clear()
        self.table.setRowCount(len(diff))
        self.table.setColumnCount(2)
        self.table.setColumnWidth(0, 400)
        self.table.setColumnWidth(1, 400)
    
        # Marque en orange dans la deuxième colonne les lignes "déplacées"
        lines_to_mark_in_second_column = set()
    
        # Compteur de différences
        num_differences = 0
    
        for i, line in enumerate(diff):
            if line.startswith('- '):
                removed_line = line[2:]
                item = QTableWidgetItem(removed_line)
                if any(line.startswith('+ ') and line[2:] == removed_line for line in diff):
                    item.setForeground(QColor("orange"))
                    lines_to_mark_in_second_column.add(removed_line)
                else:
                    item.setForeground(QColor("red"))
                self.table.setItem(i, 0, item)
                num_differences += 1
            elif line.startswith('+ '):
                added_line = line[2:]
                item = QTableWidgetItem(added_line)
                if added_line in lines_to_mark_in_second_column:
                    item.setForeground(QColor("orange"))
                    lines_to_mark_in_second_column.remove(added_line)
                else:
                    item.setForeground(QColor("limegreen"))
                self.table.setItem(i, 1, item)
                num_differences += 1
            elif line.startswith('  '):
                item1 = QTableWidgetItem(line[2:])
                item2 = QTableWidgetItem(line[2:])
                self.table.setItem(i, 0, item1)
                self.table.setItem(i, 1, item2)
    
        # Affiche un message avec le nombre de différences
        if num_differences == 0:
            QMessageBox.information(self, "Résultat de la comparaison", "Aucune différence trouvée entre les deux documents.")
        else:
            QMessageBox.information(self, "Résultat de la comparaison", f"Nombre de différences trouvées : {num_differences}")

    def showNextDifference(self):
        if self.current_difference_index < len(self.diff) - 1:
            self.current_difference_index += 1
            while self.current_difference_index < len(self.diff) and not (
                self.diff[self.current_difference_index].startswith('- ')
                or self.diff[self.current_difference_index].startswith('+ ')
            ):
                self.current_difference_index += 1
            if self.current_difference_index < len(self.diff):
                self.table.selectRow(self.current_difference_index)

    def showPreviousDifference(self):
        if self.current_difference_index > 0:
            self.current_difference_index -= 1
            while self.current_difference_index > 0 and not (
                self.diff[self.current_difference_index].startswith('- ')
                or self.diff[self.current_difference_index].startswith('+ ')
            ):
                self.current_difference_index -= 1
            if self.current_difference_index >= 0:
                self.table.selectRow(self.current_difference_index)

    def extraire_commentaires(self, pdf_file):
        doc = fitz.open(pdf_file)
        annotations_dict = {}  # Dictionnaire pour stocker les annotations

        for page_num in range(len(doc)):
            page = doc[page_num]
            annotations_on_page = []
            current_strikeout = None

            for annot in page.annots():
                annotation_info = {
                    "type": annot.type[1],
                    "content": annot.info.get("content", "").replace("\r", "\n"),  # Remplace les "\r" par des sauts de ligne
                }

                if annot.type[1] == "StrikeOut":
                    if current_strikeout:
                        annotations_on_page.append(current_strikeout)
                    current_strikeout = annotation_info
                    texte_barre = page.get_text("text", clip=annot.rect)
                    current_strikeout["texte_barre"] = texte_barre.strip()

                elif annot.type[1] == "Caret":
                    if current_strikeout:
                        # Supprimer les caractères de retour chariot du contenu du Caret
                        caret_content = annotation_info["content"].replace("\r", "\n")
                        current_strikeout["content"] = caret_content
                        annotations_on_page.append(current_strikeout)
                        current_strikeout = None
                    else:
                        annotations_on_page.append(annotation_info)

                else:
                    annotations_on_page.append(annotation_info)

            if current_strikeout:
                annotations_on_page.append(current_strikeout)

            if annotations_on_page:
                annotations_dict[page_num + 1] = annotations_on_page

        doc.close()
        return annotations_dict

    def displayOldPDFAnnotations(self):
        if self.old_pdf:
            old_pdf_annotations = self.extraire_commentaires(self.old_pdf)
    
            # Effacez d'abord le contenu précédent de la liste
            self.annotation_list.clear()

            
            for page_num, annotations_on_page in old_pdf_annotations.items():
                for i, annot_info in enumerate(annotations_on_page):
                    item_text = f"Page {page_num} - Annotation {i + 1}:\nType: {annot_info['type']}"
                    if annot_info['type'] == 'StrikeOut':
                        item_text += f"\nTexte barré: {annot_info['texte_barre']}"
                    item_text += f"\nContent: {' ' * 4}{annot_info['content']}"
            
                    # Créez une case à cocher pour chaque annotation
                    checkbox = QCheckBox()
                    checkbox.setMaximumWidth(20)  # Définir la largeur maximale de la case à cocher
                    item = QListWidgetItem()
            
                    # Creation du text_label et configuration de son contenu
                    text_label = QLabel(item_text)
                    text_label.setWordWrap(True)  # Permet au texte de se déplacer à la ligne
            
                    # Ajout du text_label à la mise en page
                    layout = QHBoxLayout()
                    layout.addWidget(checkbox)
                    layout.addWidget(text_label)
                    widget = QWidget()
                    widget.setLayout(layout)
                    item.setSizeHint(QSize(300, widget.sizeHint().height() + 10))
                    self.annotation_list.addItem(item)
                    self.annotation_list.setItemWidget(item, widget)
            
                    # Connection de la case à cocher à la fonction pour changer la couleur de fond
                    checkbox.stateChanged.connect(lambda state, item=item: self.changeBackground(item, state))

    def changeBackground(self, item, state):
        if state == Qt.Checked:
            item.setBackground(QColor(144, 238, 144))  # Couleur verte clair
        else:
            item.setBackground(QColor(255, 255, 255))

    def exportToExcel(self):
        if not self.diff:
            QMessageBox.warning(self, "Aucune Comparaison", "Veuillez d'abord comparer les fichiers PDF avant d'exporter vers Excel.")
            return
            
        data = []
    
        for i in range(self.table.rowCount()):
            old_version_item = self.table.item(i, 0)
            new_version_item = self.table.item(i, 1)
    
            old_version_text = old_version_item.text() if old_version_item else ''
            new_version_text = new_version_item.text() if new_version_item else ''
    
            # Obtient la couleur de la ligne
            color = 'red' if old_version_item and old_version_item.foreground().color() == QColor('red') else \
                    'limegreen' if new_version_item and new_version_item.foreground().color() == QColor('limegreen') else \
                    'orange' if old_version_text != new_version_text else None
    
            data.append({'Ancienne version': old_version_text, 'Nouvelle version': new_version_text, 'Couleur': color})
    
        # Crée un DataFrame à partir de la liste
        df = pd.DataFrame(data)
        
        # Demande à l'utilisateur l'emplacement où enregistrer le fichier Excel
        excel_filename, _ = QFileDialog.getSaveFileName(self, "Exporter vers Excel", "", "Excel Files (*.xlsx);;All Files (*)")
        
        if excel_filename:
            try:
                # Écrit les données dans le fichier Excel
                with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name='Résultats', index=False, startrow=1, header=False)
        
                    # Obtient la feuille de calcul créée
                    sheet = writer.sheets['Résultats']
        
                    # Crée un style pour chaque couleur
                    red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                    green_fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
                    orange_fill = PatternFill(start_color='FFFFA500', end_color='FFFFA500', fill_type='solid')
                    white_fill = PatternFill(start_color='FFFFFFFF', end_color='FFFFFFFF', fill_type='solid')
        
                    # Applique le style en fonction de la couleur dans les colonnes 'Ancienne version' et 'Nouvelle version'
                    for idx, row in df.iterrows():
                        old_version_cell = sheet.cell(row=idx + 2, column=df.columns.get_loc('Ancienne version') + 1)
                        new_version_cell = sheet.cell(row=idx + 2, column=df.columns.get_loc('Nouvelle version') + 1)
        
                        old_version_cell.fill = orange_fill if row['Couleur'] == 'orange' else red_fill if row['Couleur'] == 'red' else white_fill
                        new_version_cell.fill = orange_fill if row['Couleur'] == 'orange' else green_fill if row['Couleur'] == 'limegreen' else white_fill
        
                    # Ajoute un en-tête coloré
                    header_fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
                    for col_num, value in enumerate(['Ancienne version', 'Nouvelle version'], 1):
                        sheet.cell(row=1, column=col_num, value=value).fill = header_fill
        
                QMessageBox.information(self, "Export Excel", f"Résultats exportés avec succès vers {excel_filename}")
            except Exception as e:
                QMessageBox.critical(self, "Erreur d'exportation", f"Une erreur s'est produite lors de l'exportation vers Excel : {str(e)}")

    # Ajoutez cette fonction à votre classe PDFComparerApp
    def goToText(self):
        search_text = self.page_entry.text().strip()
    
        if not search_text:
            QMessageBox.warning(self, "Texte invalide", "Veuillez entrer un texte à rechercher.")
            return
    
        found_rows = []
        for i in range(self.table.rowCount()):
            for j in range(self.table.columnCount()):
                item = self.table.item(i, j)
                if item and search_text.lower() in item.text().lower():
                    found_rows.append(i)
                    break  # Si trouvé dans cette ligne, passer à la suivante
    
        if found_rows:
            # Sélectionnez la première occurrence trouvée
            self.table.selectRow(found_rows[0])
            # Optionnel : Faites défiler la vue vers la ligne sélectionnée
            self.table.scrollToItem(self.table.item(found_rows[0], 0))
            if len(found_rows) > 1:
                QMessageBox.information


    def showInstructions(self):
        instructions = (
            """
            <html>
            <body>
            <p><span style="font-weight: bold; text-decoration: underline;">Instructions d'utilisation de l'application "PDF Compare & Verify"</span></p>
            <ol>
                <li>Sélectionnez l'ancienne version du fichier PDF en cliquant sur le bouton "Sélectionner" à côté de "Ancienne version".</li>
                <li>Sélectionnez la nouvelle version du fichier PDF en cliquant sur le bouton "Sélectionner" à côté de "Nouvelle version".</li>
                <li>Cliquez sur le bouton "Comparer" pour comparer les deux fichiers PDF.</li>
                <li>Les différences seront affichées dans le tableau :</li>
                <ul>
                    <li>Le texte supprimé sera en <span style="color: red; font-weight: bold;">rouge</span>.</li>
                    <li>Le texte ajouté sera en <span style="color: limegreen; font-weight: bold;">vert</span>.</li>
                    <li>Les lignes de texte déplacées seront marquées en <span style="color: orange; font-weight: bold;">orange</span>.</li>
                </ul>
            </ol>
            <p>
                Utilisez les boutons "Différence précédente" et "Différence suivante" pour naviguer entre les différences mises en évidence.<br>
                <br>
                L'onglet de droite affiche les annotations. Grâce au tableau de comparaison, vous pouvez alors voir si les différences sont liées aux annotations ou non. 
                Si c'est le cas, vous pouvez indiquer que l'annotation est bien prise en compte en la cochant. 
                Elle devient alors <span style="color: limegreen; font-weight: bold;">verte</span>.
                <br>
                <br>
                Version du 13/11/2023
            </p>
            </body>
            </html>
            """
        )
    
        help_box = QMessageBox()
        help_box.setWindowTitle("Instructions d'utilisation")
        help_box.setTextFormat(Qt.RichText)
        help_box.setText(instructions)
        help_box.exec_()

def main(): 
    app = QApplication(sys.argv)
    window = PDFComparerApp()
    window.show()
    sys.exit(app.exec_())

if __name__ == "__main__":
    main()